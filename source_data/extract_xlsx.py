import openpyxl

wb = openpyxl.load_workbook('/app/source_data/aspen.xlsm', data_only=True, keep_vba=False)
print("Sheet names:")
for s in wb.sheetnames:
    print(f"  - {s}")
print("="*80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n\n{'='*80}\nSHEET: {sheet_name}  (dims: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column})\n{'='*80}")
    # Dump non-empty cells
    row_count = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=False):
        row_data = []
        for cell in row:
            if cell.value is not None:
                v = cell.value
                if isinstance(v, float):
                    v = round(v, 3)
                row_data.append(f"{cell.coordinate}={v}")
        if row_data:
            print(" | ".join(row_data))
            row_count += 1
    if row_count == 0:
        print("(empty)")
